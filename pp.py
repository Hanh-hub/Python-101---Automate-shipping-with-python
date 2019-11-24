from xlrd import open_workbook
import openpyxl
import pandas as pd
from pandas import ExcelWriter
from pandas import ExcelFile
import numpy as np
from Product import Product
from Product import Node
from Product import Queue



wb = open_workbook('input.xlsx')
sheet = wb.sheet_by_name("Sheet1")
number_of_rows = sheet.nrows
number_of_columns = sheet.ncols


shelf = Queue()

product = []
product.append(Product(None, None, None, None, None, None))

for i in range(1, number_of_rows):
    product.append(Product(sheet.cell(i, 0).value,
                           sheet.cell(i, 1).value,
                           sheet.cell(i, 2).value,
                           sheet.cell(i, 3).value,
                           sheet.cell(i, 4).value,
                           sheet.cell(i, 5).value))
for i in range(number_of_rows):
    shelf.add(Node(product[i]))

def update():
    counter = shelf.getSize()
    node = shelf.head
    node = node.next
    product_update = []
    list_a = []
    df =[]
    product_update.append(Product(None, None, None, None, None, None))
    list_a.append(None)
    df.append(None)
    xlsfile = 'output.xlsx'

    writer = pd.ExcelWriter(xlsfile, engine='xlsxwriter')
    for i in range(1, counter):
        product_update.append(Product(node.data.sku, node.data.length, node.data.width, node.data.height,
                                      node.data.frag, node.data.weight))
        node = node.next
        list_a.append([product_update[i].sku,product_update[i].length,
                       product_update[i].width,product_update[i].height,
                       product_update[i].frag,product_update[i].weight])
        df.append([pd.DataFrame(list_a[i])])


    header=['Sku','length','width','height','fragile','weight']
    #header=['a','b']
    df1 = pd.DataFrame(header)
        # df1.transpose()
    df1 = df1.transpose()

    df1.to_excel(writer, sheet_name="Sheet1",
                     startrow=0, startcol=0, header=False, index=False)
    for i in range(0,counter):
        df[i] = pd.DataFrame(list_a[i])
        # df1.transpose()
        df[i] = df[i].transpose()

        df[i].to_excel(writer, sheet_name="Sheet1",
                     startrow=i, startcol=0, header=False, index=False)

    writer.save()
    print('successfully export to excel')



loop= True
while loop:
    print(' [1] to ship an item \n'
            ' [2] to display the inventory\n'
            ' [3] to sort items new to old \n'
            ' [4] to update the inventory to spreadsheet\n'
            ' [5] to quit the program \n')
    opt = input("Choose one option: ")
    print()
    if opt == '1':
        ans=True
        while ans:
            print()

            sku=input('What item do you want to ship? \n (Press 0 to go back to menu): ')

            if sku=='0':
                ans=False
            else:
                shelf.shipping(sku)
        print('-----------------------------------')

    elif opt == '2':
        shelf.print_inventory()
        print('-----------------------------------')

    elif opt == '3':
        shelf.traversal_tail_to_head()
        print('-----------------------------------')
    elif opt is '4':
        update()
        print()
    elif opt == '5':
        print("Program ended")
        exit()
    else:
        print("Invalid input. Please try again! :)")
        print()
